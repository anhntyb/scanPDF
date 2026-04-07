from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .models import ExtractionResult


def export_result(result: ExtractionResult, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["template_id", result.template_id or "unknown"])
    ws.append(["pdf", str(result.pdf.path)])
    ws.append(["pages", result.pdf.pages])
    ws.append(["image_only", result.pdf.is_image_only])

    raw = wb.create_sheet("raw_ocr")
    raw.append(["page", "confidence", "text"])
    for page in result.pages:
        raw.append([page.page_number, page.confidence, page.text])

    parsed = wb.create_sheet("parsed_data")
    if result.parsed_rows:
        preferred = [
            "source_file",
            "template_id",
            "page_number",
            "row_no",
            "code",
            "full_name",
            "bank_account",
            "amount",
            "currency",
            "unit_name",
            "title",
            "month",
            "year",
            "bank_name",
            "raw_line",
        ]
        headers = [h for h in preferred if h in result.parsed_rows[0]]
        extras = [h for h in result.parsed_rows[0].keys() if h not in headers]
        headers.extend(extras)
        parsed.append(headers)
        for row in result.parsed_rows:
            parsed.append([row.get(h) for h in headers])
    else:
        parsed.append(["status", "No parsed rows yet"])

    for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        parsed.column_dimensions[column].width = 18
    parsed.column_dimensions["F"].width = 28
    parsed.column_dimensions["O"].width = 90

    flags = wb.create_sheet("review_flags")
    flags.append(["warning"])
    for warning in result.warnings:
        flags.append([warning])

    wb.save(output)
    return output
